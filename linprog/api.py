from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uvicorn
import json
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from datetime import datetime
import requests
import uuid
from process_dnf import BoardProcessor, CHANNEL_COUNT_FIELDS, process_dnf_requirements_core
from optimize import optimize_card_selection_core
import sys
# 添加路径以导入query_sim
sys.path.insert(0, os.path.dirname(__file__))
from query_sim import query_all_sim_machines, generate_output_format

# 配置环境变量
API_KEY = os.getenv('API_KEY', 'sk-zzvwbcaxoss3')
FILE_SERVER_URL = os.getenv('FILE_SERVER_URL', 'https://demo.langcore.cn')


# API_KEY = os.getenv('API_KEY', 'sk-6zvekr4931xm')
# FILE_SERVER_URL = os.getenv('FILE_SERVER_URL', 'http://10.120.120.6:3008')


app = FastAPI(
    title="板卡选型优化 API",
    description="基于线性规划的板卡最优采购方案计算+excel生成",
    version="1.0.0"
)

# 配置 CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 通道类型：直接使用 process_dnf.py 中的 CHANNEL_COUNT_FIELDS（39个字段）
CHANNEL_TYPES = CHANNEL_COUNT_FIELDS

# 通道类型数量（39个）
CHANNEL_COUNT = len(CHANNEL_COUNT_FIELDS)

# 请求模型


class CardInfo(BaseModel):
    matrix_channel_count: List[int] = Field(
        ..., description=f"{CHANNEL_COUNT}个元素的数组，表示各通道类型的数量（对应CHANNEL_COUNT_FIELDS的顺序）")
    model: str = Field(..., description="板卡型号")
    price_cny: int = Field(..., description="板卡价格（人民币）")
    id: str = Field(..., description="板卡唯一标识ID")
    original: Optional[List[str]] = Field(
        None, description="原始需求描述列表（可选，来自process_dnf输出）")


class OptimizationRequest(BaseModel):
    """优化请求模型，直接使用 process_dnf 输出格式"""
    linprog_input_data: List[Dict[str, Any]] = Field(
        ..., description="process_dnf输出的板卡数据（包含id, matrix_channel_count, model, price_cny, original）")
    linprog_requiremnets: List[int] = Field(
        ..., description=f"process_dnf输出的需求数组（{CHANNEL_COUNT}个元素）")

# 响应模型


class OptimizedCard(BaseModel):
    model: str
    quantity: int
    unit_price: int
    total_price: int
    id: str
    original: Optional[List[str]] = Field(
        None, description="原始需求描述列表（可选，来自process_dnf输出）")


class ChannelSatisfaction(BaseModel):
    channel_type: str
    required: int
    satisfied: int
    status: str


class FeasibilityCheck(BaseModel):
    channel_type: str
    required: int
    available_total: int
    max_single_card: int
    status: str


class OptimizationResponse(BaseModel):
    success: bool
    message: str
    total_cards: int
    requirements_summary: List[dict]
    feasibility_checks: List[FeasibilityCheck]
    optimized_solution: Optional[List[OptimizedCard]] = None
    total_cost: Optional[int] = None
    channel_satisfaction: Optional[List[ChannelSatisfaction]] = None
    unsatisfied_requirements: List[dict] = []


@app.get("/")
async def root():
    """API 根路径"""
    return {
        "message": "板卡选型优化 API",
        "version": "1.0.0",
        "docs": "/docs"
    }


@app.get("/channel-types")
async def get_channel_types():
    """获取所有通道类型列表"""
    return {
        "channel_types": CHANNEL_TYPES,
        "count": len(CHANNEL_TYPES)
    }


@app.post("/optimize", response_model=OptimizationResponse)
async def optimize_card_selection(request: OptimizationRequest):
    """
    板卡选型优化接口

    直接使用 process_dnf 输出格式：
    - **linprog_input_data**: process_dnf输出的板卡数据数组（包含id, matrix_channel_count, model, price_cny, original）
    - **linprog_requiremnets**: process_dnf输出的需求数组（{CHANNEL_COUNT}个元素）

    返回最优采购方案，包括总成本和每种板卡的采购数量
    """
    try:
        # 调用核心优化函数
        result = optimize_card_selection_core(
            linprog_input_data=request.linprog_input_data,
            linprog_requiremnets=request.linprog_requiremnets
        )

        # 将字典结果转换为 Pydantic 模型
        # 转换 feasibility_checks
        feasibility_checks = [
            FeasibilityCheck(**fc) for fc in result.get('feasibility_checks', [])
        ]

        # 转换 optimized_solution
        optimized_solution = None
        if result.get('optimized_solution'):
            optimized_solution = [
                OptimizedCard(**card) for card in result['optimized_solution']
            ]

        # 转换 channel_satisfaction
        channel_satisfaction = None
        if result.get('channel_satisfaction'):
            channel_satisfaction = [
                ChannelSatisfaction(**cs) for cs in result['channel_satisfaction']
            ]

        return OptimizationResponse(
            success=result['success'],
            message=result['message'],
            total_cards=result['total_cards'],
            requirements_summary=result['requirements_summary'],
            feasibility_checks=feasibility_checks,
            optimized_solution=optimized_solution,
            total_cost=result.get('total_cost'),
            channel_satisfaction=channel_satisfaction,
            unsatisfied_requirements=result.get('unsatisfied_requirements', [])
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ================= process_dnf 接口 =================

class RequirementItem(BaseModel):
    """需求项模型"""
    id: Optional[str] = Field(None, description="需求ID")
    original: str = Field(..., description="原始需求描述")
    DNF: str = Field(..., description="DNF逻辑表达式")


class ProcessDNFRequest(BaseModel):
    """process_dnf 请求模型"""
    require: List[RequirementItem] = Field(...,
                                           description="需求列表，每个需求包含 original 和 DNF 字段")


class ProcessDNFResponse(BaseModel):
    """process_dnf 响应模型"""
    success: bool
    message: str
    timestamp: str
    linprog_input_data: List[Dict[str, Any]]
    linprog_requiremnets: List[int]
    matched_boards: List[Dict[str, Any]]
    total_candidates: int
    total_matches: int
    processing_info: Dict[str, Any]
    unsatisfied_requirements: List[Dict[str, Any]] = Field(
        default_factory=list, description="无法处理的需求（DNF为空或没有百分百匹配的板卡）")


@app.post("/process-dnf", response_model=ProcessDNFResponse)
async def process_dnf_requirements(request: ProcessDNFRequest):
    """
    处理DNF逻辑表达式，查询数据库，生成板卡匹配结果

    - **require**: 需求列表，每个需求包含：
      - **original**: 原始需求描述
      - **DNF**: DNF逻辑表达式（如: "AD_channel_count_single_ended ≥ 16 and DA_channel_count ≥ 16"）
      - **id**: 可选的需求ID

    返回匹配的板卡数据和线性规划输入数据
    """
    try:
        # 构建输入数据格式（转换为字典列表）
        require_list = [
            {
                'id': req.id if req.id else f"req_{idx}_{uuid.uuid4().hex[:8]}",
                'original': req.original,
                'DNF': req.DNF
            }
            for idx, req in enumerate(request.require)
        ]

        # 调用核心处理函数
        output_data = process_dnf_requirements_core(require=require_list)

        return ProcessDNFResponse(
            success=True,
            message="处理成功",
            timestamp=output_data['timestamp'],
            linprog_input_data=output_data['linprog_input_data'],
            linprog_requiremnets=output_data['linprog_requiremnets'],
            matched_boards=output_data['matched_boards'],
            total_candidates=output_data['total_candidates'],
            total_matches=output_data['total_matches'],
            processing_info=output_data['processing_info'],
            unsatisfied_requirements=output_data.get('unsatisfied_requirements', [])
        )

    except Exception as e:
        import traceback
        error_detail = f"处理失败: {str(e)}\n{traceback.format_exc()}"
        raise HTTPException(status_code=500, detail=error_detail)


# ================= query_sim 接口 =================

class SimRequirementItem(BaseModel):
    """仿真机需求项模型"""
    original: str = Field(..., description="原始需求描述")
    attribute: Dict[str, Any] = Field(..., description="需求属性字典")


class QuerySimRequest(BaseModel):
    """查询仿真机请求模型"""
    require: List[SimRequirementItem] = Field(..., description="需求列表，每个需求包含 original 和 attribute 字段")


class QuerySimResponse(BaseModel):
    """查询仿真机响应模型"""
    success: bool
    message: str
    result_id: Dict[str, Any] = Field(..., description="最佳匹配仿真机的详细信息")
    sim_raw_data: List[Dict[str, Any]] = Field(..., description="最佳匹配仿真机的原始数据")
    sim_pick_list: List[Dict[str, Any]] = Field(..., description="按需求分类的仿真机列表")


@app.post("/query-sim", response_model=QuerySimResponse)
async def query_sim_machines(request: QuerySimRequest):
    """
    查询满足条件的仿真机

    - **require**: 需求列表，每个需求包含：
      - **original**: 原始需求描述（如："CPU：不低于八核Intel CoreI9，主频不低于 4.0GHz处理器"）
      - **attribute**: 需求属性字典（如：{"cpu_cores": 8, "cpu_frequency_value": 4, ...}）

    返回最佳匹配的仿真机及其详细信息
    """
    try:
        # 构建输入数据格式（转换为字典列表）
        require_list = [
            {
                'original': req.original,
                'attribute': req.attribute
            }
            for req in request.require
        ]

        # 查询所有仿真机
        all_machines = query_all_sim_machines()
        
        # 生成输出格式
        output_data = generate_output_format(all_machines, require_list)

        return QuerySimResponse(
            success=True,
            message="查询成功",
            result_id=output_data.get('result_id', {}),
            sim_raw_data=output_data.get('sim_raw_data', []),
            sim_pick_list=output_data.get('sim_pick_list', [])
        )

    except Exception as e:
        import traceback
        error_detail = f"查询失败: {str(e)}\n{traceback.format_exc()}"
        raise HTTPException(status_code=500, detail=error_detail)


# ================= Excel 生成功能 =================

def generate_combined_excel(json_data: Dict[str, Any], template_path: str, output_path: str):
    """
    Generate combined Excel file with two sheets: quotation sheet and requirement matching preview

    Args:
        json_data: JSON data dictionary
        template_path: Excel template file path
        output_path: Output Excel file path

    Returns:
        str: Output file path
        float: Total amount
    """

    # 1. 先生成报价单（基于模板）
    wb = load_workbook(template_path)

    # 删除不需要的工作表，只保留第一个
    if len(wb.sheetnames) > 1:
        for sheet_name in wb.sheetnames[1:]:
            wb.remove(wb[sheet_name])

    # 重命名第一个工作表为"报价单"
    ws1 = wb.active
    ws1.title = "报价单"

    # 定义安全写入单元格的函数（默认上下左右居中）
    def safe_cell_write(ws, row, col, value):
        """安全地写入单元格值，避免合并单元格错误，并统一设置居中"""
        try:
            cell = ws.cell(row=row, column=col)
            # 检查单元格类型
            if cell.__class__.__name__ == 'MergedCell':
                # 如果是合并单元格，跳过写入
                return None
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            return cell
        except Exception as e:
            print(f"警告：无法写入单元格 ({row}, {col}): {e}")
            return None

    # 自动识别模板表头位置
    def detect_header_positions(ws):
        """在前50行内寻找包含关键表头的一行，返回该行号及列映射"""
        required_headers = ["型号", "数量", "单价", "总价"]
        optional_headers = ["序号", "税率", "应税", "总金额"]
        header_row = None
        header_map = {}

        max_rows = min(ws.max_row, 50)
        max_cols = min(ws.max_column, 30)

        for r in range(1, max_rows + 1):
            current_map = {}
            for c in range(1, max_cols + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str):
                    text = v.strip()
                else:
                    text = v
                if text in required_headers + optional_headers:
                    current_map[text] = c
            # 至少匹配3个必需表头才认为找到
            if sum(1 for h in required_headers if h in current_map) >= 3:
                header_row = r
                header_map = current_map
                break

        return header_row, header_map

    # 处理报价单数据（自动识别新模板表头）
    header_row, header_map = detect_header_positions(ws1)
    start_row = (header_row + 1) if header_row else 13
    # 兼容多种格式：card、array
    result_items = json_data.get('card', json_data.get('array', []))
    raw_sim_items = json_data.get('raw_sim', [])

    # ========== 根据ID进行去重和合并 ==========
    def merge_items_by_id(items):
        """根据板卡ID对数据进行去重和合并（同一张卡满足多个需求）"""
        merged_dict = {}

        for item in items:
            # 兼容新旧两种格式
            if 'each_obj' in item:
                each_obj = item.get('each_obj', {})
            else:
                each_obj = item

            item_id = each_obj.get('id', '')
            if not item_id:
                # 如果没有ID，直接跳过
                continue

            if item_id not in merged_dict:
                # 第一次遇到这个ID，初始化
                merged_dict[item_id] = {
                    'id': item_id,
                    'originals': [],
                    'match_degrees': [],
                    'reasons': [],
                    'type': each_obj.get('type', ''),
                    'model': each_obj.get('model', ''),
                    'description': each_obj.get('description', each_obj.get('brief_description', '')),
                    'manufacturer': each_obj.get('manufacturer', ''),
                    'price_cny': each_obj.get('price_cny', 0),
                    'quantity': each_obj.get('quantity', 1),  # 取第一条记录的数量
                    'details': []
                }

            # 收集需要合并的字段（同一张卡的多个需求）
            # original 字段可能是列表或字符串格式
            original = each_obj.get('original', [])
            if isinstance(original, list) and original:
                # 如果是列表，展开并添加到 originals 中，确保所有元素都是字符串
                for item in original:
                    if isinstance(item, str):
                        merged_dict[item_id]['originals'].append(item)
                    elif isinstance(item, list):
                        # 如果是嵌套列表，递归展开（虽然不应该出现，但为了安全）
                        merged_dict[item_id]['originals'].extend([str(x) for x in item])
                    else:
                        # 其他类型转换为字符串
                        merged_dict[item_id]['originals'].append(str(item))
            elif isinstance(original, str) and original:
                # 如果是字符串，直接添加到 originals 中
                merged_dict[item_id]['originals'].append(original)

            match_degree = each_obj.get(
                'match_degree', each_obj.get('score', ''))
            if match_degree:
                merged_dict[item_id]['match_degrees'].append(str(match_degree))

            reason = each_obj.get('reason', '')
            if reason:
                merged_dict[item_id]['reasons'].append(reason)

            # 处理details（如果存在）
            if 'details' in each_obj and each_obj['details']:
                merged_dict[item_id]['details'].extend(each_obj['details'])

        # 将合并后的数据转换为列表
        merged_list = []
        for item_id, data in merged_dict.items():
            # 确保 originals 中的所有元素都是字符串
            originals_str = [str(item) if not isinstance(item, str) else item for item in data['originals']]
            
            merged_item = {
                'id': data['id'],
                'original': '\n'.join(originals_str),
                'match_degree': '\n'.join(data['match_degrees']),
                'reason': '\n'.join(data['reasons']),
                'type': data['type'],
                'model': data['model'],
                'description': data['description'],
                'manufacturer': data['manufacturer'],
                'price_cny': data['price_cny'],
                'quantity': data['quantity'],  # 不累加，保持原数量
                'details': data['details'] if data['details'] else None
            }
            merged_list.append(merged_item)

        return merged_list

    # 对 card 数据根据ID进行合并
    result_items = merge_items_by_id(result_items)

    current_row = start_row
    total_amount = 0

    # 处理 card 板卡数据（已合并）——适配新模板：只写“序号/型号/数量/单价/总价”
    running_index = 1
    for each_obj in result_items:
        item_id = each_obj.get('id', '')
        price_cny = each_obj.get('price_cny', 0)
        quantity = each_obj.get('quantity', 1)

        # 计算总金额
        total_amount_cny = price_cny * quantity

        # 转数字
        try:
            if isinstance(price_cny, str):
                price_cny = float(price_cny.replace('￥', '').replace(',', ''))
            if isinstance(total_amount_cny, str):
                total_amount_cny = float(
                    total_amount_cny.replace('￥', '').replace(',', ''))
        except (ValueError, AttributeError):
            price_cny = 0
            total_amount_cny = 0

        # 累加总价
        total_amount += total_amount_cny

        # 列索引（优先表头映射，回退旧模板列位）
        col_index = {
            '序号': header_map.get('序号', 2),
            '型号': header_map.get('型号', 7),
            '数量': header_map.get('数量', 11),
            '单价': header_map.get('单价', 10),
            '总价': header_map.get('总价', 12)
        }

        if col_index['序号']:
            safe_cell_write(ws1, current_row, col_index['序号'], running_index)
        if col_index['型号']:
            safe_cell_write(ws1, current_row, col_index['型号'], each_obj.get(
                'model', each_obj.get('type', '')))
        if col_index['数量']:
            safe_cell_write(ws1, current_row, col_index['数量'], quantity)
        if col_index['单价']:
            safe_cell_write(ws1, current_row, col_index['单价'], price_cny)
        if col_index['总价']:
            safe_cell_write(ws1, current_row,
                            col_index['总价'], total_amount_cny)

        ws1.row_dimensions[current_row].height = None
        current_row += 1
        running_index += 1

    # 处理 raw_sim 机箱数据（同样写到新模板列）
    for item in raw_sim_items:
        sim_id = item.get('id', '')
        price_cny = item.get('price_cny', 0)
        quantity = item.get('quantity', 1)

        # 计算总金额
        total_amount_cny = price_cny * quantity

        # 转换价格为数字格式
        try:
            if isinstance(price_cny, str):
                price_cny = float(price_cny.replace('￥', '').replace(',', ''))
            if isinstance(total_amount_cny, str):
                total_amount_cny = float(
                    total_amount_cny.replace('￥', '').replace(',', ''))
        except (ValueError, AttributeError):
            price_cny = 0
            total_amount_cny = 0

        # 累加总价
        total_amount += total_amount_cny

        # 列索引
        col_index = {
            '序号': header_map.get('序号', 2),
            '型号': header_map.get('型号', 7),
            '数量': header_map.get('数量', 11),
            '单价': header_map.get('单价', 10),
            '总价': header_map.get('总价', 12)
        }

        if col_index['序号']:
            safe_cell_write(ws1, current_row, col_index['序号'], running_index)
        if col_index['型号']:
            safe_cell_write(ws1, current_row, col_index['型号'], item.get(
                'model', item.get('type', '')))
        if col_index['数量']:
            safe_cell_write(ws1, current_row, col_index['数量'], quantity)
        if col_index['单价']:
            safe_cell_write(ws1, current_row, col_index['单价'], price_cny)
        if col_index['总价']:
            safe_cell_write(ws1, current_row,
                            col_index['总价'], total_amount_cny)

        # 设置行高为自动
        ws1.row_dimensions[current_row].height = None
        current_row += 1

    # 找到F列值为"小计"和"总计"的行，并填入总金额到L列
    subtotal_row = None
    total_row = None

    # 遍历所有行，寻找"小计"和"总计"（适配任意列）
    for row in range(1, ws1.max_row + 1):
        try:
            found_text = None
            for col in range(1, ws1.max_column + 1):
                val = ws1.cell(row=row, column=col).value
                if val in ("小计", "总计"):
                    found_text = val
                    break
            if found_text == "小计":
                subtotal_row = row
                write_col = header_map.get('总金额') or header_map.get('总价') or 12
                safe_cell_write(ws1, row, write_col, total_amount)
                ws1.row_dimensions[row].height = None
            elif found_text == "总计":
                total_row = row
                write_col = header_map.get('总金额') or header_map.get('总价') or 12
                safe_cell_write(ws1, row, write_col, total_amount)
                ws1.row_dimensions[row].height = None
        except Exception as e:
            print(f"警告：处理第{row}行时出错: {e}")
            continue

    print(f"找到小计行: {subtotal_row}, 总计行: {total_row}")
    print(f"填入总金额: {total_amount}")

    # 2. 创建需求匹配预览工作表
    ws2 = wb.create_sheet("需求匹配预览")

    # 设置表头
    headers = [
        "序号", "原始需求", "规格型号", "技术参数", "制造商",
        "数量", "单价(元)", "小计(元)"
    ]

    # 美化表头
    header_fill = PatternFill(start_color="4F81BD",
                              end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thick'))

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 设置表头行高
    ws2.row_dimensions[1].height = 25

    # 从第2行开始写入数据
    current_row = 2
    total_amount = 0

    # 建立需求点与设备的映射关系（合并后的数据作为整体）
    demand_device_mapping = []

    for each_obj in result_items:
        # 合并后的数据，original已经是换行连接的多个需求
        original_text = each_obj.get('original', '')
        if original_text:
            # 将合并后的数据作为一个整体添加
            demand_device_mapping.append({
                'demand': original_text,  # 保持换行符
                'reason': each_obj.get('reason', ''),  # 保持换行符
                'device': each_obj
            })

    # 处理 raw_sim 机箱数据
    for item in raw_sim_items:
        # 机箱使用category作为需求
        demand = item.get('category', '机箱&CPU控制器')
        demand_device_mapping.append({
            'demand': demand,
            'reason': '',  # 机箱一般不需要匹配原因
            'device': item
        })

    # 写入数据行
    for idx, mapping in enumerate(demand_device_mapping, 1):
        demand = mapping['demand']
        reason = mapping['reason']
        device = mapping['device']

        # 写入数据行
        ws2.cell(row=current_row, column=1, value=idx)
        ws2.cell(row=current_row, column=2, value=demand)
        ws2.cell(row=current_row, column=3, value=device.get('model', ''))
        # 兼容字段：description -> detailed_description -> brief_description
        tech_desc = device.get('description', device.get(
            'detailed_description', device.get('brief_description', '')))
        ws2.cell(row=current_row, column=4, value=tech_desc)
        ws2.cell(row=current_row, column=5,
                 value=device.get('manufacturer', ''))
        ws2.cell(row=current_row, column=6, value=device.get('quantity', 1))

        # 计算单价
        price_cny = device.get('price_cny', 0)
        if isinstance(price_cny, str):
            price_cny = float(price_cny.replace('￥', '').replace(',', ''))
        ws2.cell(row=current_row, column=7, value=price_cny)

        # 小计 - 如果没有total_amount_cny，则计算
        total_amount_cny = device.get('total_amount_cny', 0)
        if isinstance(total_amount_cny, str):
            total_amount_cny = float(
                total_amount_cny.replace('￥', '').replace(',', ''))
        elif total_amount_cny == 0:
            # 如果没有提供total_amount_cny，则计算
            total_amount_cny = price_cny * device.get('quantity', 1)
        ws2.cell(row=current_row, column=8, value=total_amount_cny)

        # 美化数据行样式
        data_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # 隔行换色
        if current_row % 2 == 0:
            row_fill = PatternFill(start_color="F2F2F2",
                                   end_color="F2F2F2", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFF",
                                   end_color="FFFFFF", fill_type="solid")

        data_font = Font(size=11)

        # 为每一列设置样式
        for col in range(1, 9):
            cell = ws2.cell(row=current_row, column=col)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = data_border

            # 全部统一上下左右居中
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 对可能的多行文本开启换行，但保持居中
        ws2.cell(row=current_row, column=2).alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        ws2.cell(row=current_row, column=4).alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

        # 设置行高为自动
        ws2.row_dimensions[current_row].height = None
        current_row += 1

    # 优化列宽，对ws1主报价单表自适应
    for col in range(1, ws1.max_column + 1):
        maxlen = 8
        for row in range(1, ws1.max_row + 1):
            val = ws1.cell(row=row, column=col).value
            if val is None:
                val = ""
            else:
                val = str(val)
            # 中文字符宽度更大，按比例增大
            try:
                chinese_count = sum(1 for ch in val if ord(ch) > 127)
                eff_len = len(val) + chinese_count  # 中文计宽加权
            except:
                eff_len = len(val)
            if eff_len > maxlen:
                maxlen = eff_len
        ws1.column_dimensions[get_column_letter(
            col)].width = min(max(8, maxlen*1.8), 35)

    # ws2需求匹配预览表还是用固定列宽
    column_widths = [8, 35, 18, 35, 15, 10, 15, 15]
    column_names = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    for i, (width, col_name) in enumerate(zip(column_widths, column_names)):
        ws2.column_dimensions[col_name].width = width

    # 保存文件
    wb.save(output_path)
    return output_path, total_amount


def generate_excel_from_json_string(json_string: str, template_content: bytes) -> tuple:
    """
    Generate Excel file from JSON string and template file content

    Args:
        json_string: JSON data as string
        template_content: Excel template file content as bytes

    Returns:
        tuple: (output_file_path, total_amount)
    """
    try:
        # 解析JSON数据
        json_data = json.loads(json_string)

        # 使用项目temp目录而不是系统临时目录
        project_temp_dir = os.path.join(os.path.dirname(__file__), "temp")
        os.makedirs(project_temp_dir, exist_ok=True)

        # 生成唯一的文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        temp_template_path = os.path.join(
            project_temp_dir, f"template_{timestamp}.xlsx")
        output_path = os.path.join(
            project_temp_dir, f"output_{timestamp}.xlsx")

        # 写入模板文件
        with open(temp_template_path, 'wb') as f:
            f.write(template_content)

        try:
            # 生成Excel文件
            result_path, total_amount = generate_combined_excel(
                json_data, temp_template_path, output_path)
            return result_path, total_amount
        finally:
            # 清理临时模板文件
            try:
                if os.path.exists(temp_template_path):
                    os.unlink(temp_template_path)
            except:
                pass  # 忽略删除错误

    except json.JSONDecodeError as e:
        raise ValueError(f"无效的JSON数据: {str(e)}")
    except Exception as e:
        raise ValueError(f"生成Excel文件时出错: {str(e)}")


async def upload_file_to_server(file_path: str, token: str = None) -> Dict[str, Any]:
    """
    上传文件到服务器

    Args:
        file_path: 要上传的文件路径
        token: 认证token

    Returns:
        dict: 上传结果
    """
    try:
        headers = {}
        if token:
            headers['Authorization'] = f'Bearer {token}'

        with open(file_path, 'rb') as f:
            files = {'file': f}
            response = requests.post(
                f'{FILE_SERVER_URL}/api/file',
                files=files,
                headers=headers,
                timeout=30
            )

        result = {
            "success": response.status_code == 200,
            "status_code": response.status_code
        }

        if response.status_code == 200 and response.text:
            try:
                # 解析返回的JSON字符串
                response_data = json.loads(response.text)
                if response_data.get("status") == "success" and "data" in response_data:
                    file_id = response_data["data"].get("fileId")
                    if file_id:
                        # 构造完整的文件URL
                        full_url = f"{FILE_SERVER_URL}/api/file/{file_id}"
                        result["file_url"] = full_url
                        result["file_id"] = file_id
                        result["response"] = f"文件上传成功，访问地址: {full_url}"
                    else:
                        result["response"] = response.text
                else:
                    result["response"] = response.text
            except json.JSONDecodeError:
                # 如果不是JSON格式，直接返回原始文本
                result["response"] = response.text
        else:
            result["response"] = response.text if response.text else "上传成功"

        return result
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


@app.post("/generate-excel")
async def generate_excel(json_data: Dict[str, Any], token: str = API_KEY):
    """
    生成Excel文件并自动上传

    - **json_data**: 包含配置数据的JSON对象
    - **token**: 认证token（默认为 API_KEY ）
    - **returns**: 生成和上传的结果
    """
    try:
        # 获取默认模板文件路径
        template_path = os.path.join(os.path.dirname(__file__), "空白输出模板.xlsx")

        # 验证模板文件是否存在
        if not os.path.exists(template_path):
            raise HTTPException(
                status_code=404,
                detail=f"默认模板文件不存在: {template_path}"
            )

        # 读取模板文件内容
        with open(template_path, 'rb') as template_file:
            template_content = template_file.read()

        # 将JSON数据转换为字符串
        json_string = json.dumps(json_data)

        # 生成Excel文件
        output_path, total_amount = generate_excel_from_json_string(
            json_string, template_content)

        # 上传文件到服务器
        upload_result = await upload_file_to_server(output_path, token)

        # 清理临时文件
        try:
            if os.path.exists(output_path):
                os.unlink(output_path)
        except:
            pass  # 忽略清理错误

        # 返回结果
        response_data = {
            "message": "Excel文件生成并上传成功",
            "total_amount": total_amount,
            "upload_result": upload_result
        }

        # 如果上传成功且有file_url，添加到响应中
        if upload_result.get("success") and "file_url" in upload_result:
            response_data["file_url"] = upload_result["file_url"]
            response_data["file_id"] = upload_result.get("file_id")

        return response_data

    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"无效的JSON数据: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"处理失败: {str(e)}")


@app.get("/health")
async def health_check():
    """健康检查"""
    return {"status": "healthy"}


if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)
